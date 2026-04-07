# %%
import ast
from typing import Any

import ipdb
import numpy as np
import pandas as pd
from fran.configs.mnemonics import Mnemonics
from fran.preprocessing.helpers import (
    create_dataset_stats_artifacts,
    infer_dataset_stats_window,
    postprocess_artifacts_missing,
)
from fran.utils.folder_names import (
    folder_names_from_plan,
)
from fran.utils.string_works import is_excel_None
from label_analysis.totalseg import TotalSegmenterLabels
from utilz.fileio import load_yaml
from utilz.stringz import ast_literal_eval

tr = ipdb.set_trace


from openpyxl import load_workbook
from utilz.helpers import Path, get_train_valid_test_lists_from_json, os, pp, tr

KEYS_STR_TO_LIST = ("spacing", "patch_size", "expand_by", "ignore_labels")
# HACK: this may bug out later
REMAPPING_DICT_OR_LIST = {
    "remapping_source": "dict",
    "remapping_lbd": "dict",
    "remapping_whole": "dict",
    "remapping_train": "dict",
    "remapping_imported": "dict",
}


def parse_excel_remapping(remapping) -> list:
    remapping = ast_literal_eval(remapping)
    if isinstance(remapping, str):
        remapping = remapping.split(",")
        remapping = [ast_literal_eval(rems) for rems in remapping]
    if not isinstance(remapping, list | tuple):
        remapping = [remapping]
    return remapping


def parse_excel_datasources(datasources: str) -> list:
    datasources = datasources.replace(" ", "").split(",")
    return datasources


def cases_in_folder(fldr) -> int:
    fldr = Path(fldr)
    if not fldr.exists():
        return 0
    img_fldr = fldr / ("images")
    cases = list(img_fldr.glob("*"))
    n_cases = len(cases)






    return n_cases


def confirm_plan_analyzed(project, plan):

    n_cases = len(project)
    folders = folder_names_from_plan(project, plan)
    existing_src_fldr = folders["data_folder_source"]
    cases_in_src_folder = cases_in_folder(existing_src_fldr)
    src_fldr_full = n_cases == cases_in_src_folder

    mode = plan.get("mode")
    if mode == "lbd":
        existing_final_fldr = folders["data_folder_lbd"]
    elif mode in ["patch", "pbd"]:
        existing_final_fldr = folders["data_folder_pbd"]
    elif mode == "whole":
        existing_final_fldr = folders["data_folder_whole"]
    elif mode == "source" or mode == "sourcepbd":
        existing_final_fldr = folders["data_folder_source"]
    else:
        raise NotImplementedError(f"Unknown mode: {mode}")

    cases_in_final_folder = cases_in_folder(existing_final_fldr)
    final_fldr_full = n_cases == cases_in_final_folder
    return {"src_fldr_full": src_fldr_full, "final_fldr_full": final_fldr_full}


def _to_py(obj) -> Any:
    """Recursively convert numpy scalars to Python scalars and cast 1.0 -> 1."""
    # numpy scalar -> Python scalar
    if isinstance(obj, np.generic):
        obj = obj.item()

    # containers
    if isinstance(obj, list):
        return [_to_py(x) for x in obj]
    if isinstance(obj, tuple):
        return tuple(_to_py(x) for x in obj)
    if isinstance(obj, dict):
        return {_to_py(k): _to_py(v) for k, v in obj.items()}

    # floats that are integer-valued -> int
    if isinstance(obj, float) and obj.is_integer():
        return int(obj)

    return obj


def labels_from_remapping(remapping_in):
    def _inner(remapping):
        if is_excel_None(remapping) or remapping == "":
            return 0
        if isinstance(remapping, str) and "TSL" in remapping:
            TSL = TotalSegmenterLabels()
            attr = remapping.replace(" ", "").replace("TSL.", "").split(":")[1]
            dest = getattr(TSL, attr)
        elif isinstance(remapping, dict):
            dest = list(remapping.values())
        if 0 not in dest:
            dest = [0] + dest
        # return dest
        dest_total = int(max(dest)) + 1
        return dest_total
        # fall through to remapping/global if parsing fails

    if remapping_in is None or remapping_in == "" or remapping_in == "nan":
        return 0
        # return int(max(dest)) + 1
    if not isinstance(remapping_in, list | tuple):
        remapping = [remapping_in]
    else:
        remapping = remapping_in

    remappings_out = []
    for rem in remapping:
        labels_all = _inner(rem)
        remappings_out.append(labels_all)
    labels_all = max(remappings_out)
    return labels_all



def parse_nested_remapping(plan, key, as_list=False, as_dict=False):
    def _parse_single_remapping(remapping):
        if isinstance(remapping, str) and "TSL" in remapping:
            src, dest = remapping.replace(" ", "").split(":")
            src = src.split(".")[1]
            dest = dest.split(".")[1]
            TSL = TotalSegmenterLabels()
            remapping = TSL.create_remapping(
                src, dest, as_list=as_list, as_dict=as_dict
            )
        elif isinstance(remapping, dict) and as_dict == True:
            remapping = remapping
        elif (
            isinstance(remapping, list) and as_list == True and len(remapping) == 2
        ):  # its in correct format
            remapping = remapping
        elif remapping is None:
            remapping = None
        else:
            raise NotImplementedError
        return remapping

    assert as_list or as_dict, "Either list mode or dict mode should be true"

    if key not in plan.keys():
        remapping = None
    else:
        remapping = plan[key]
    if not remapping:
        return
    remapping = parse_excel_remapping(remapping)

    datasources = plan["datasources"]
    datasources = parse_excel_datasources(datasources)
    assert len(datasources) == len(remapping), (
        "For each datasource, a unique remapping is required, it can be None."
    )
    remappings_out = []
    for ds, remapping in zip(datasources, remapping):
        remapping_out = _parse_single_remapping(remapping)
        remappings_out.append(remapping_out)
    return remappings_out


def parse_excel_dict(dici, keys_str_to_list) -> dict:
    if not isinstance(dici, dict):
        return _to_py(dici)
    for key, value in list(dici.items()):
        if isinstance(value, dict):
            dici[key] = parse_excel_dict(value, keys_str_to_list)
            continue
        if is_excel_None(value):
            dici[key] = None
            continue
        if key in keys_str_to_list and value is not None:
            try:
                value = ast_literal_eval(value)
            except Exception:
                pass
        dici[key] = _to_py(value)

    dici = maybe_add_patch_size(dici)
    dici = maybe_add_src_dims(dici)
    return dici


def maybe_add_patch_size(plan):
    # if "patch_size" in plan.keys():
    #     return plan
    if "patch_dim0" and "patch_dim1" in plan.keys():
        plan["patch_size"] = make_patch_size(plan["patch_dim0"], plan["patch_dim1"])
    return plan


def maybe_add_src_dims(dataset_params):
    if "src_dim0" and "src_dim1" in dataset_params.keys():
        dataset_params["src_dims"] = make_patch_size(
            dataset_params["src_dim0"], dataset_params["src_dim1"]
        )
    return dataset_params


BOOL_ROWS = "patch_based,one_cycles,heavy,deep_supervision,self_attention,fake_tumours,square_in_union,apply_activation"


def check_bool(row):
    if row["var_name"] in BOOL_ROWS.split(","):
        row["manual_value"] = bool(row["manual_value"])
    return row


def make_patch_size(patch_dim0, patch_dim1):
    patch_size = [
        patch_dim0,
    ] * 2 + [
        patch_dim1,
    ]
    return patch_size


def get_imagelists_from_config(project, fold, patch_based, dim0, dim1):
    json_fname = project.validation_folds_filename
    if patch_based == False:
        folder_name = project.stage1_folder / "{0}_{1}_{1}/images".format(
            dim0, dim1, dim1
        )
        train_list, valid_list, _ = get_train_valid_test_lists_from_json(
            project_title=project.project_title,
            fold=fold,
            json_fname=json_fname,
            image_folder=folder_name,
            ext=".pt",
        )
        # Note: This print statement would need verbose parameter passed to this function
        print("Retrieved whole image datasets from folder: {}".format(folder_name))
    else:
        train_list, valid_list, _ = get_train_valid_test_lists_from_json(
            project_title=project.project_title,
            fold=fold,
            image_folder=project.stage1_folder / "cropped/images_pt/images",
            json_fname=json_fname,
            ext=".pt",
        )

    return train_list, valid_list


def load_config_from_worksheet(settingsfilename, sheet_name, engine="pd"):
    """
    Reads a sheet row-by-row
    """

    if engine == "pd":
        df = pd.read_excel(settingsfilename, sheet_name=sheet_name)
    elif engine == "openpyxl":  # THIS IS NOT PROPERLY IMPLEMENTED AND HAS BUGS
        wb = load_workbook(settingsfilename)
        sheet = wb[sheet_name]
        df = pd.DataFrame(sheet.values)

    else:
        raise NotImplementedError
    config = {}
    for row in df.iterrows():
        rr = row[1]
        rr = check_bool(rr)
        var_type = rr["tune_type"]
        key = rr["var_name"]
        if sheet_name == "transform_factors":
            val = parse_excel_cell(rr["manual_value"])
            prob = rr["manual_p"]
            config.update({key: [val, prob]})
        else:
            val = parse_excel_cell(rr["manual_value"])
            config.update({key: parse_excel_cell(val)})
    return config


class ConfigMaker:
    def __init__(
        self,
        project,
    ):
        self.project = project
        configuration_mnemonic = project.global_properties["mnemonic"]
        configuration_filename = self.resolve_configuration_filename()
        plans = pd.read_excel(
            configuration_filename,
            sheet_name="plans",
            index_col="id",
            keep_default_na=False,
            na_values=["TRUE", "FALSE", ""],
        )
        configuration_mnemonic_standardized = Mnemonics.match(configuration_mnemonic)
        self.plans = plans.loc[plans["mnemonic"] == configuration_mnemonic_standardized]
        self.plans = self.plans.drop(columns=["mnemonic"])
        self.plans.insert(0, "plan_id", self.plans.index)
        self.plans = self.plans.set_index("plan_id", drop=False)
        configs = load_config_from_workbook(configuration_filename)
        self.configs = parse_excel_dict(configs, KEYS_STR_TO_LIST)

    def setup(
        self,
        plan_train: int,
        plan_valid: int = None,
        plan_test: int = None,
        verbose=True,
    ):  # , plan_valid=None):
        # by default plan_valid is a fixed plan regardless of train_plan and is set in dataset_params
        # plan_valid essentially only uses the folder of said plan, and patch_size is kept same as plan_train
        if plan_valid is None:
            plan_valid = self.plans.loc[plan_train]["plan_valid"]
        if plan_test is None:
            plan_test = self.plans.loc[plan_train]["plan_test"]

        if is_excel_None(plan_valid):
            plan_valid = plan_train
        if is_excel_None(plan_test):
            plan_test = plan_train

        self._set_active_plans(plan_train, plan_valid, plan_test)
        self.add_dataset_props()

    def resolve_configuration_filename(self):

        common_vars_filename = os.environ["FRAN_CONF"] + "/config.yaml"
        common_paths = load_yaml(common_vars_filename)
        configurations_folder = Path(common_paths["configurations_folder"])
        return configurations_folder / ("experiment_configs.xlsx")

    def add_dataset_props(self):
        props = [
            "intensity_clip_range",
            "mean_fg",
            "std_fg",
            "mean_dataset_clipped",
            "std_dataset_clipped",
        ]
        for prop in props:
            try:
                self.configs["dataset_params"][prop] = self.project.global_properties[
                    prop
                ]
            except:
                self.configs["dataset_params"][prop] = None

    # CODE: depcrecated add_out_channels as per folder labels_out are being computed in preprocessing

    def maybe_merge_source_plan(self, plan_key="plan_train"):
        """Merge source plan into the specified plan
        Args:
            config: Configuration dictionary
            plan_key: Key of the plan to merge into ('plan_train' or 'plan_valid')
        Returns:
            Updated config dictionary
        """
        # Retrieve the main plan and source plan from the config dictionary
        main_plan = self.configs[plan_key]
        src_plan_key = main_plan.get("source_plan")
        if is_excel_None(src_plan_key):
            return

        # Ensure the source plan exists in the config before proceeding
        else:
            # Access the source plan
            src_plan_key = src_plan_key.replace(" ", "")
            src_plan_k, src_plan_mode = src_plan_key.split(",")
            src_plan_k = ast_literal_eval(src_plan_k)
            source_plan = self.plans.loc[src_plan_k]
            source_plan = dict(source_plan)
            self.configs["plan_source"] = source_plan
            # source_plan = config.get(src_plan_key, {})

            # Iterate over the source plan keys and add any missing keys to the main plan
            for key in source_plan:
                main_plan_val = main_plan.get(key)
                if is_excel_None(main_plan_val):
                    main_plan[key] = source_plan[key]
            # self.configs[plan_key] = main_plan

    def _set_plan(self, plan_id, suffix: str):
        assert suffix in [
            "train",
            "valid",
            "test",
        ], "suffix must be either 'train', 'valid' or 'test'"
        """Helper function to set a plan configuration
        Args:
            plan_num: Plan number from dataset_params
        """
        plan_id = plan_id
        plan_selected = self.plans.loc[plan_id]
        plan_selected = dict(plan_selected)
        samples_per_file = plan_selected["samples_per_file"]
        plan_selected["samples_per_file"] = (
            int(samples_per_file) if not is_excel_None(samples_per_file) else 1
        )
        plan_key = "plan_" + suffix
        self.configs[plan_key] = plan_selected
        self.maybe_merge_source_plan(plan_key)
        if is_excel_None(plan_selected["expand_by"]):
            plan_selected["expand_by"] = 0

        self.configs[plan_key] = parse_excel_dict(plan_selected, KEYS_STR_TO_LIST)
        self.configs[plan_key]["plan_name"] = plan_id

        for key, value in REMAPPING_DICT_OR_LIST.items():
            # reg = load_registry()
            # org_val = self.configs[plan_key][key]
            # self.configs[plan_key][key+"_code"] = remapping_conv(reg=reg, key="remapping",val=org_val)
            if key in self.configs[plan_key].keys():
                self.configs[plan_key][key] = parse_nested_remapping(
                    plan=self.configs[plan_key],
                    key=key,
                    as_dict=True if value == "dict" else False,
                    as_list=True if value == "list" else False,
                )

    def _set_active_plans(
        self, plan_train: int = None, plan_valid: int = None, plan_test: int = None
    ):
        # if plan_train == None:
        #     plan_train = self.configs["dataset_params"]["plan_train"]
        # if plan_valid == None:
        #     plan_valid = self.configs["dataset_params"]["plan_valid"]
        self._set_plan(plan_train, "train")
        self._set_plan(plan_valid, "valid")
        self._set_plan(plan_test, "test")
        self.validate_plans()
        self.configs["plan_valid"]["patch_size"] = self.configs["plan_train"][
            "patch_size"
        ]
        self.configs["plan_test"]["patch_size"] = self.configs["plan_train"][
            "patch_size"
        ]

    def validate_plans(self):
        for remp_key in ["remapping_source", "remapping_lbd"]:
            for plan_name in "plan_valid", "plan_test":
                assert (
                    self.configs[plan_name][remp_key]
                    == self.configs["plan_train"][remp_key]
                ), f"{plan_name} {remp_key} is not the same as plan_train {remp_key}"

    def add_preprocess_status(self):
        """Add preprocessing status column to plans dataframe"""

        preprocess = []
        n_cases = len(self.project)

        for plan_id in self.plans.index:
            self.setup(plan_id, False)
            conf = self.configs
            plan = conf["plan_train"]

            detailed_status = confirm_plan_analyzed(self.project, plan)
            src_fldr_full = detailed_status["src_fldr_full"]
            final_fldr_full = detailed_status["final_fldr_full"]

            if all([src_fldr_full, final_fldr_full]):
                status = "both"
            elif any([src_fldr_full, final_fldr_full]):
                status = "one"
            else:
                status = "none"

            preprocess.append(status)

        self.plans["preprocessed"] = preprocess

    def create_plan_postproc_artifacts(
        self, gif=True, label_stats=True, gif_window="abdomen"
    ):
        plan = self.configs["plan_train"]
        mode = plan["mode"]
        ff = folder_names_from_plan(self.project, plan)
        data_folder_key = "data_folder_" + mode
        data_folder = ff[data_folder_key]
        missing = postprocess_artifacts_missing(data_folder)
        create_gif = gif == True and missing["gif"] == True
        create_label_stats = label_stats == True and missing["label_stats"] == True
        if create_gif == False and create_label_stats == False:
            print("No artifacts to create")
            return missing
        create_dataset_stats_artifacts(
            output_folder=data_folder,
            gif=create_gif,
            label_stats=create_label_stats,
            gif_window=gif_window,
        )
        #     folder_key = "data_folder_whole"
        #     generator_cls = FixedSizeDataGenerator
        #     generator = generator_cls(
        #         project=self.project,
        #         plan=plan,
        #         data_folder=folders["data_folder_source"],
        #         output_folder=folders[folder_key],
        #     )
        # elif mode in ["source", "sourcepbd"]:
        #     from fran.preprocessing.fixed_spacing import NiftiToTorchDataGenerator
        #
        #     folder_key = "data_folder_source"
        #     generator_cls = NiftiToTorchDataGenerator
        #     generator = generator_cls(
        #         project=self.project,
        #         plan=plan,
        #         data_folder=self.project.raw_data_folder,
        #         output_folder=folders[folder_key],
        #     )
        # else:
        #     raise NotImplementedError(f"Unknown mode: {mode}")
        #
        # output_folder = Path(folders[folder_key])
        # stats_folder = output_folder / "dataset_stats"
        # missing_gif = not (stats_folder / "snapshot.gif").exists()
        # missing_label_stats = not (stats_folder / "lesion_stats.csv").exists()
        # missing_labels = not (output_folder / "labels_all.json").exists()
        # missing_props = not (output_folder / "resampled_dataset_properties.json").exists()
        #
        # generator.store_gifs = missing_gif
        # generator.store_label_stats = missing_label_stats
        #
        # if missing_labels or missing_props or missing_gif or missing_label_stats:
        #     generator.run_postprocess_only()
        # return {
        #     "mode": mode,
        #     "output_folder": output_folder,
        #     "missing_gif": missing_gif,
        #     "missing_label_stats": missing_label_stats,
        #     "missing_labels": missing_labels,
        #     "missing_props": missing_props,
        # }
        #
        #


def load_config_from_workbook(settingsfilename) -> dict:
    wb = load_workbook(settingsfilename)
    sheets = wb.sheetnames
    configs_dict = {}
    for sheet in sheets:
        if sheet.lower() == "plans":
            # Read all plans at once and spread them into plan1, plan2, ...
            pass
        else:
            # Old behavior for all other sheets
            configs_dict[sheet] = load_config_from_worksheet(settingsfilename, sheet)

    return configs_dict
    # sheets.remove("metadata")


def load_metadata(settingsfilename):
    df = pd.read_excel(settingsfilename, sheet_name="metadata", index_col=None)
    return df


def parse_excel_cell(cell_val):
    if isinstance(cell_val, (int, float)):
        return _to_py(cell_val)
    if isinstance(cell_val, str):
        if any(p in cell_val for p in ("[", "{", "(")):
            try:
                return _to_py(ast.literal_eval(cell_val))
            except:
                return cell_val
        else:
            try:
                return _to_py(ast.literal_eval(cell_val))
            except:
                return cell_val
    return _to_py(cell_val)


def normalize_logging_payload(value):
    if isinstance(value, dict):
        return {kk: normalize_logging_payload(vv) for kk, vv in value.items()}
    if isinstance(value, list):
        return [normalize_logging_payload(vv) for vv in value]
    if isinstance(value, tuple):
        return [normalize_logging_payload(vv) for vv in value]
    return parse_excel_cell(value)


def parse_neptune_dict(dic: dict):
    return normalize_logging_payload(dic)


# %%
if __name__ == "__main__":
# SECTION:-------------------- setup-------------------------------------------------------------------------------------- <CR> <CR> <CR> <CR> <CR>

    # set_autoreload()
    from fran.managers import Project

    P = Project(project_title="test")
    P = Project(project_title="pancreas")
    P = Project(project_title="kidneys")
    P = Project(project_title="totalseg")

# %%
    P.global_properties
    C = ConfigMaker(P)
    C.setup(2)
    pp(C.configs["plan_train"])
    pp(C.configs["plan_valid"])
    C.configs["plan_train"].keys()
    C.configs["plan_train"]["labels_all_lbd"]
    plan = C.configs["plan_train"]
    pp(plan["spacing"])
    mode = plan["mode"]
    C.create_plan_postproc_artifacts()
# %%

    main_plan = C.configs["plan_train"]
    src_plan_key = main_plan.get("source_plan")

    # Ensure the source plan exists in the config before proceeding
    if src_plan_key:
        # Access the source plan
        src_plan_key = src_plan_key.replace(" ", "")
        src_plan_k, src_plan_mode = src_plan_key.split(",")
        src_plan_k = ast_literal_eval(src_plan_k)
        source_plan = C.plans.loc[src_plan_k]
        source_plan = dict(source_plan)
        C.configs["plan_source"] = source_plan
        # source_plan = config.get(src_plan_key, {})

        # Iterate over the source plan keys and add any missing keys to the main plan
# %%
        for key in source_plan:
            print(key)
            if key == "spacing":
                tr()
            main_plan_val = main_plan.get(key)
            if main_plan_val is None:
                main_plan[key] = source_plan[key]
# %%
    C.plans["mode"]
# %%
    df = C.plans
# %%
    conf = C.configj
    pp(conf["dataset_params"])
    pp(conf["plan_train"])
    conf["plan_train"]["imported_folder"]
    conf["plan_train"]["remapping_imported"]
# %%

    existing_fldr = folder_names_from_plan(P, plan)["data_folder_source"]
    img_fldr = existing_fldr / ("images")
    len(list(img_fldr.glob("*"))) == len(project)
# %%

    lbd_subfolder = folder_names_from_plan(project, plan)["data_folder_lbd"]
    lbd_img_fldr = Path(lbd_subfolder) / ("images")
    len(list(lbd_img_fldr.glob("*")))
# %%
    df = C.plans
    row = df.iloc[2]
    plan = row.to_dict()

# %%

    conf["dataset_params"]["src_dims"] = make_patch_size(
        conf["dataset_params"]["src_dim0"], conf["dataset_params"]["src_dim1"]
    )
    conf["plan_train"]["patch_size"] = make_patch_size(
        conf["plan_train"]["patch_dim0"], conf["plan_train"]["patch_dim1"]
    )
# %%
    plan = C.configs["plan_train"]
# %%
    global_props = C.project.global_properties
    global_props["labels_all"]
    rmt = plan.get("remapping_train")
    rmt = ast_literal_eval(rmt)

    rms = plan.get("remapping_source")

    rms = ast_literal_eval(rms)
    rml = plan.get("remapping_lbd")

    rml = ast_literal_eval(rml)
    rmi = plan.get("remapping_imported")

    rmi = ast_literal_eval(rmi)
    rmw = plan.get("remapping_whole")
    rmw = ast_literal_eval(rmw)
# %%
    mode = plan.get("mode")
    labels_all_train = labels_from_remapping(rmt)
    # elif (mode == "source" or mode == "whole") and rms:
    labels_all_source = labels_from_remapping(rms)
    # else: labels_all = global_props["labels_all"]
    # elif mode == "lbd" and rml:
    labels_all_lbd = labels_from_remapping(rml)

    # elif mode == "whole" and rmw:
    labels_all_whole = labels_from_remapping(rmw)
    labels_all_datasources = global_props["labels_all"]
# %%
    dici = {}
    dici["labels_all_train"] = labels_all_train
    dici["labels_all_source"] = labels_all_source
    if labels_all_source == 0:
        labels_all_source = sum(global_props["labels_all"])
    dici["labels_all_lbd"] = labels_all_lbd
    dici["labels_all_whole"] = labels_all_whole
    dici["labels_all_datasources"] = labels_all_datasources

# %%

    train = True
    plan_num = 3
    plan_name = "plan" + str(plan_num)
    plan_selected = C.plans.loc[plan_name]
    plan_selected = dict(plan_selected)
    samples_per_file = plan_selected["samples_per_file"]
    plan_selected["samples_per_file"] = (
        int(samples_per_file) if not is_excel_None(samples_per_file) else 1
    )
    plan_key = "plan_train" if train else "plan_valid"
    C.configs[plan_key] = plan_selected
    plan = plan_selected.copy()
    pp(plan)
# %%
    # C.maybe_merge_source_plan(plan_key)
    C.configs[plan_key] = parse_excel_dict(plan_selected)

    pp(C.configs[plan_key])
    C.config 





