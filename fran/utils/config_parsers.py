# %%
import ast
import sys
import warnings

import ipdb
import pandas as pd
from fastcore.basics import store_attr
from label_analysis.totalseg import TotalSegmenterLabels
from utilz.fileio import load_yaml
from utilz.string import ast_literal_eval
from fran.managers.datasource import MNEMONICS
import numpy as np

tr = ipdb.set_trace

if not sys.executable == "":  # workaround for slicer as it does not load ray tune
    from ray import tune

from openpyxl import load_workbook
from utilz.helpers import *


def is_excel_None(input):
    if not input:
        return True
    input = str(input)
    if input == "nan":
        return True
    else:
        return False



def _to_py(obj):
    """Recursively convert numpy scalars to Python scalars and cast 1.0 -> 1."""
    # numpy scalar -> Python scalar
    if isinstance(obj, np.generic):
        obj = obj.item()

    # containers
    if isinstance(obj, list):
        return [_to_py(x) for x in obj]
    if isinstance(obj, tuple):
        return tuple(_to_py(x) for x in obj)
    if isinstance(obj, dict):
        return { _to_py(k): _to_py(v) for k, v in obj.items() }

    # floats that are integer-valued -> int
    if isinstance(obj, float) and obj.is_integer():
        return int(obj)

    return obj
def labels_from_remapping(remapping):
    def _inner(remapping):
            if remapping is None or remapping == "" or remapping == "nan":
                return None
            if isinstance(remapping, pd.Series):
                try: remapping = ast.literal_eval(remapping.item())
                except Exception: pass

            # TSL.* or explicit (src, dest)
            try:
                if isinstance(remapping, (list, tuple)) and len(remapping) == 2:
                    dest = remapping[1]
                    # return int(max(dest)) + 1
                elif isinstance(remapping, str) and "TSL" in remapping:
                    TSL = TotalSegmenterLabels()
                    attr = remapping.replace("TSL.","").split(",")[1]
                    # use the class-ids list to keep rule: max(dest)+1
                    dest = getattr(TSL, attr)
                    # return int(max(dest)) + 1
                if 0 not in dest:
                    dest = [0] + dest
                return dest
            except Exception as e:
                print(e)
                return None
            # fall through to remapping/global if parsing fails
    # --- 1) remapping_train ---
    # remapping = plan.get("remapping_train")
        # pandas cell stored as string/Series -> try literal_eval
    labels_all = _inner(remapping)
    return set(labels_all)

#CODE: delete the below if code is not breaking  (see #13)
def compute_out_labels(plan: dict, global_props: dict | None = None) -> list:
    """
    Priority:
      1) plan['remapping_train']  -> infer mapping, return max(dest)+1
      2) plan['remapping']        -> if [src, dest] or (src, dest) use max(dest)+1; if dict use max(values)+1
      3) global_props['labels_all'] -> len + 1
      4) default to 2
    """
    # --- 1) remapping_train ---
    rmt = plan.get("remapping_train")
    rms = plan.get("remapping_source")
    rml= plan.get("remapping_lbd")
    mode = plan.get("mode")
    # if all([rmt,rms,rml]) is False:
    #     return global_props["labels_all"]
    if rmt: labels_all  = labels_from_remapping(rmt)
    elif (mode == "source" or mode=="whole") and rms:
         labels_all = labels_from_remapping(rms)
        # else: labels_all = global_props["labels_all"]
    elif mode=="lbd" and rml:
         labels_all = labels_from_remapping(rml)
        # else: return global_props["labels_all"]
    else:
        labels_all = global_props["labels_all"]
    return labels_all
    # --- 2) remapping ---
    remap = plan.get("remapping")
    if isinstance(remap, (list, tuple)) and len(remap) == 2:
        dest = remap[1]
        return int(max(dest)) + 1
    if isinstance(remap, dict) and remap:
        return int(max(remap.values())) + 1

    # --- 3) global fallback ---

    labels_all=[]
    for ds in global_props['datasources']:
        labs =         ds['labels']
        labels_all.extend(labs)
        labels_all = set(labels_all)
        print("Unique labels in all datasets:", labels_all)
        fg = len(labels_all)
        return fg+1

    # if global_props and "labels_all" in global_props:
    #     oc = len(global_props["labels_all"]) + 1
    #     return max(2, oc)

    # --- 4) last resort ---
    warnings.warn("Could not infer out_channels; defaulting to 2 (BG+FG).")
    return 2

def create_remapping(plan,key, as_list=False,as_dict=False):
        assert  as_list or as_dict, "Either list mode or dict mode should be true"
        if key not in  plan.keys():
            remapping = None
        else:
            remapping = plan[key]

        if isinstance(remapping, str) and "TSL" in remapping:
            src,dest = remapping.split(",")
            src = src.split(".")[1]
            dest = dest.split(".")[1]
            TSL  = TotalSegmenterLabels()
            remapping = TSL.create_remapping(src,dest,as_list=as_list,as_dict=as_dict)
        elif isinstance(remapping, dict) and as_dict ==True:
            remapping =remapping
        elif isinstance(remapping, list) and as_list ==True and len(remapping)==2: # its in correct format
            remapping  = remapping
        elif remapping is None :
            remapping = None
        else: raise NotImplementedError
        return remapping


def parse_excel_dict(dici):
    """Recursively parse an Excel plan, handling nested dictionaries

    Args:
        plan: Dictionary containing plan configuration, possibly with nested dictionaries
    Returns:
        Parsed plan with proper types for all values
    """
    if not isinstance(dici, dict):
        return dici

    # keys_maybe_nan = "fg_indices_exclude", "lm_groups", "datasources", "cache_rate", "ce_weight", "remapping_sitk","remapping_train","remapping_imported"
    keys_str_to_list = "spacing", "patch_size"

    for key, value in dici.items():
        # Handle nested dictionaries recursively
        if isinstance(value, dict):
            dici[key] = parse_excel_dict(value)
            continue

        # Handle None values
        if is_excel_None(value):
            dici[key] = None
            continue

        # Handle string to list conversion
        if key in keys_str_to_list and value is not None:
            try:
                dici[key] = ast_literal_eval(value)
            except (ValueError, SyntaxError, TypeError):
                # Keep original value if conversion fails
                continue

    dici = maybe_add_patch_size(dici)
    return dici

def parse_excel_dict(dici):
    if not isinstance(dici, dict):
        return _to_py(dici)

    keys_str_to_list = ("spacing", "patch_size")

    for key, value in list(dici.items()):
        if isinstance(value, dict):
            dici[key] = parse_excel_dict(value)
            continue
        if is_excel_None(value):
            dici[key] = None
            continue
        if key in keys_str_to_list and value is not None:
            try:
                value = ast_literal_eval(value)
            except Exception:
                pass
        dici[key] = _to_py(value)

    dici = maybe_add_patch_size(dici)
    return dici

def maybe_add_patch_size(plan):
    if "patch_size" in plan.keys():
        return plan
    if "patch_dim0" and "patch_dim1" in plan.keys():
        plan["patch_size"] = make_patch_size(plan["patch_dim0"], plan["patch_dim1"])
    return plan



BOOL_ROWS = "patch_based,one_cycles,heavy,deep_supervision,self_attention,fake_tumours,square_in_union,apply_activation"


def check_bool(row):
    if row["var_name"] in BOOL_ROWS.split(","):
        row["manual_value"] = bool(row["manual_value"])
    return row


def make_patch_size(patch_dim0, patch_dim1):
    patch_size = [
        patch_dim0,
    ] * 2 + [
        patch_dim1,
    ]
    return patch_size

#
# def out_channels_from_TSL(remapping_train):
#     TSL = TotalSegmenterLabels()
#     attrib = remapping_train.split(".")[1]
#     labels = getattr(TSL, attrib)
#     labels = set(labels)
#     out_ch = len(labels)
#     return out_ch


def remapping_from_remapping_train(remapping_train):
    if is_excel_None(remapping_train):
        return None
    elif "TSL" in remapping_train:
        TSL = TotalSegmenterLabels()
        TSL_attr = remapping_train.split(".")[1]
        orig_labels = TSL.all
        final_labels = getattr(TSL, TSL_attr)
        return orig_labels, final_labels
    elif isinstance(remapping_train, tuple) or isinstance(remapping_train, list):
        return remapping_train
    else:
        raise NotImplementedError

#
# def out_channels_from_global_properties(global_properties):
#     try:
#         out_ch = len(global_properties["labels_all"]) + 1
#         if out_ch < 2:
#             warnings.warn(
#                 "Out channel set at {0} by labels_all. It is being reset at 2 as minimum (1 BG, 1 FG)".format(
#                     out_ch
#                 )
#             )
#             out_ch = 2
#         return out_ch
#
#     except KeyError as er:
#         print("*" * 20)
#         print("Warning: Key {} not is in project.global_properties ".format(er))
#         print(
#             "Training will breakdown unless projectwide properties are set first. \nAlternatively set 'out_channels' key in config['model_params']  "
#         )
#         return None
#
#
# def out_channels_from_dict_or_cell(remapping_train):
#     if is_excel_None(remapping_train):
#         return None
#     if "TSL" in remapping_train:
#         out_channels = out_channels_from_TSL(remapping_train)
#         return out_channels
#     if isinstance(remapping_train, pd.core.series.Series):
#             remapping_train = ast.literal_eval(remapping_train.item())
#     if isinstance(remapping_train, tuple) or isinstance(remapping_train, list):
#         dest_labels = remapping_train[1]
#         out_channels = max(dest_labels) + 1
#     else:
#         raise NotImplementedError
#     return out_channels
#

def get_imagelists_from_config(project, fold, patch_based, dim0, dim1):
    json_fname = project.validation_folds_filename
    if patch_based == False:
        folder_name = project.stage1_folder / "{0}_{1}_{1}/images".format(
            dim0, dim1, dim1
        )
        train_list, valid_list, _ = get_train_valid_test_lists_from_json(
            project_title=project.project_title,
            fold=fold,
            json_fname=json_fname,
            image_folder=folder_name,
            ext=".pt",
        )
        print("Retrieved whole image datasets from folder: {}".format(folder_name))
    else:
        train_list, valid_list, _ = get_train_valid_test_lists_from_json(
            project_title=project.project_title,
            fold=fold,
            image_folder=project.stage1_folder / "cropped/images_pt/images",
            json_fname=json_fname,
            ext=".pt",
        )

    return train_list, valid_list


def resolve_tune_fnc(tune_type: str):
    if "_" in tune_type:
        return getattr(tune, tune_type.split("_")[0])
    else:
        return getattr(tune, tune_type)


def load_config_from_worksheet(settingsfilename, sheet_name, raytune, engine="pd"):
    """
    Reads a sheet row-by-row
    """

    if engine == "pd":
        df = pd.read_excel(settingsfilename, sheet_name=sheet_name)
    elif engine == "openpyxl":  # THIS IS NOT PROPERLY IMPLEMENTED AND HAS BUGS
        wb = load_workbook(settingsfilename)
        sheet = wb[sheet_name]
        df = pd.DataFrame(sheet.values)

    else:
        raise NotImplementedError
    config = {}
    for row in df.iterrows():
        rr = row[1]
        rr = check_bool(rr)
        var_type = rr["tune_type"]
        key = rr["var_name"]
        if sheet_name == "transform_factors":
            if raytune == False or rr["tune"] == False:
                val = parse_excel_cell(rr["manual_value"])
                prob = rr["manual_p"]
            else:
                if rr["tune_type"] == "double_range":
                    val_lower, val_upper = parse_excel_cell(rr["tune_value"])
                    val_lower = tune.uniform(lower=val_lower[0], upper=val_lower[1])
                    val_upper = tune.uniform(lower=val_upper[0], upper=val_upper[1])
                    val = [val_lower, val_upper]

                    prob = parse_excel_cell(rr["tune_p"])
                    prob = tune.uniform(lower=prob[0], upper=prob[1])
            config.update({key: [val, prob]})
        else:
            if raytune == False or rr["tune"] == False:
                val = parse_excel_cell(rr["manual_value"])
                config.update({key: parse_excel_cell(val)})
            else:
                if "spec" in var_type:
                    pass
                else:
                    val = parse_excel_cell(rr["tune_value"])
                    if "_" in var_type:
                        tr()
                        vals = parse_excel_cell(rr["tune_value"])
                        var_type.split("_")
                        tune_fnc = resolve_tune_fnc(var_type)
                        val_sample = [tune_fnc(val[0], val[1]) for val in vals]
                    elif (
                        var_type == "randint"
                        or var_type == "loguniform"
                        or var_type == "uniform"
                    ):
                        tune_fnc = resolve_tune_fnc(var_type)
                        val_sample = tune_fnc(val[0], val[1])
                    elif rr["tune_type"] == "double_range":
                        val_lower, val_upper = parse_excel_cell(rr["tune_value"])
                        val_lower = tune.uniform(lower=val_lower[0], upper=val_lower[1])
                        val_upper = tune.uniform(lower=val_upper[0], upper=val_upper[1])
                        val_sample = [val_lower, val_upper]

                    elif rr["tune_type"] == "choice":
                        val_sample = tune.choice(val)
                    else:
                        tune_fnc = resolve_tune_fnc(var_type)
                        if tune_fnc.__name__[0] == "q":
                            val.append(rr["quant"])
                        val_sample = tune_fnc(*val)
                    config.update({key: val_sample})
    return config


class ConfigMaker:
    def __init__(
        self,
        project,
        configuration_filename=None,
        raytune=False,
    ):
        store_attr()
        configuration_mnemonic = project.global_properties["mnemonic"]
        configuration_filename = self.resolve_configuration_filename(
            configuration_filename, configuration_mnemonic
        )
        self.plans = pd.read_excel(configuration_filename, sheet_name="plans",index_col="id")
        self.configs = load_config_from_workbook(configuration_filename, raytune)
        self.configs = parse_excel_dict(self.configs)
        if not "mom_low" in self.configs["model_params"].keys() and raytune == True:
            config = {
                "mom_low": tune.sample_from(
                    lambda spec: np.random.uniform(0.6, 0.9100)
                ),
                "mom_high": tune.sample_from(
                    lambda spec: np.minimum(
                        0.99,
                        spec.config.model_params.mom_low
                        + np.random.uniform(low=0.05, high=0.35),
                    )
                ),
            }
            self.configs["model_params"].update(config)

    def setup(self ,
        plan_train=None,
        plan_valid=None):
        self._set_active_plans(plan_train, plan_valid)
        self.add_output_labels()
        self.add_out_channels()
        self.add_dataset_props()


    def add_output_labels(self):
        out_labels= compute_out_labels(self.configs["plan_train"],self.project.global_properties)
        out_labels = set(out_labels)
        out_labels.update([0])
        print("labels output by Dataloaders, including background: ", out_labels)
        self.configs['plan_train']["labels_all"] = out_labels
        print("-" * 20)
        # self.config[plan]["labels_all"] = labels_all


    def resolve_configuration_filename(
        self, configuration_filename, configuration_mnemonic
    ):

        if configuration_filename:
            return configuration_filename
        assert (
            configuration_filename or configuration_mnemonic
        ), "Provide either a configuration filename or a configuration mnemonic"

        common_vars_filename = os.environ["FRAN_COMMON_PATHS"]
        common_paths = load_yaml(common_vars_filename)
        configurations_folder = Path(common_paths["configurations_folder"])
        if configuration_mnemonic:
            assert (
                configuration_mnemonic in MNEMONICS
            ), "Please provide a valid mnemonic from the list {}".format(MNEMONICS)
        if configuration_mnemonic == "liver" or configuration_mnemonic == "lits":
            return configurations_folder / ("experiment_configs_liver.xlsx")
        elif configuration_mnemonic == "lungs":
            return configurations_folder / ("experiment_configs_lungs.xlsx")
        elif configuration_mnemonic == "nodes":
            return configurations_folder / ("experiment_configs_nodes.xlsx")
        elif configuration_mnemonic == "totalseg":
            return configurations_folder / ("experiment_configs_totalseg.xlsx")


    def add_dataset_props(self):
        props = [
            "intensity_clip_range",
            "mean_fg",
            "std_fg",
            "mean_dataset_clipped",
            "std_dataset_clipped",
        ]
        for prop in props:
            try:
                self.configs["dataset_params"][prop] = self.project.global_properties[
                    prop
                ]
            except:
                self.configs["dataset_params"][prop] = None

    def add_out_channels(self):
        # out_ch = out_channels_from_dict_or_cell(
        #     self.config["plan_train"].get("remapping_train")
        # )
        # if not out_ch:
        #     out_ch = out_channels_from_global_properties(self.project.global_properties)
        out_ch = len(self.configs["plan_train"]["labels_all"])
        self.configs["model_params"]["out_channels"] = out_ch
        print("Out channels set to {}".format(out_ch))
        print("-" * 20)


    def maybe_merge_source_plan(self, plan_key="plan_train"):
        """Merge source plan into the specified plan
        Args:
            config: Configuration dictionary
            plan_key: Key of the plan to merge into ('plan_train' or 'plan_valid')
        Returns:
            Updated config dictionary
        """
        # Retrieve the main plan and source plan from the config dictionary
        main_plan = self.configs[plan_key]
        src_plan_key = main_plan.get("source_plan")

        # Ensure the source plan exists in the config before proceeding
        if src_plan_key:
            # Access the source plan
            source_plan  = self.plans.loc[src_plan_key]
            source_plan = dict(source_plan)
            # source_plan = config.get(src_plan_key, {})

            # Iterate over the source plan keys and add any missing keys to the main plan
            for key in source_plan:
                if key not in main_plan:
                    main_plan[key] = source_plan[key]

            

    def _set_plan(self, plan_num, train: bool=True):
        """Helper function to set a plan configuration
        Args:
            plan_num: Plan number from dataset_params
            train: Boolean - True for training plan, False for validation plan
        """
        plan_name = "plan" + str(plan_num)
        plan_selected = self.plans.loc[plan_name]
        plan_selected = dict(plan_selected)
        samples_per_file = plan_selected["samples_per_file"]
        plan_selected["samples_per_file"] = int(samples_per_file) if not is_excel_None(samples_per_file) else 1
        plan_key = "plan_train" if train else "plan_valid"
        self.configs[plan_key] = plan_selected
         # self.maybe_merge_source_plan(plan_key)
        self.configs[plan_key] = parse_excel_dict(plan_selected)
        self.configs[plan_key]["plan_name"] = plan_name
        self.configs[plan_key]["remapping_train"] = remapping_from_remapping_train(
            self.configs[plan_key]["remapping_train"]
        )
        self.configs[plan_key]["remapping"] = create_remapping(self.configs[plan_key],key="remapping",as_list=True)
        if "remapping_imported" in self.configs[plan_key].keys():
            self.configs[plan_key]["remapping_imported"] = create_remapping(self.configs[plan_key],key="remapping_imported",as_list=True)


    def _set_active_plans(self, plan_train: int=None, plan_valid:int =None):
        if plan_train == None:
            plan_train = self.configs["dataset_params"]["plan_train"]
        if plan_valid == None:
            plan_valid = self.configs["dataset_params"]["plan_valid"]
        self._set_plan(plan_train, True)
        self._set_plan(plan_valid, False)

       

def load_config_from_workbook(settingsfilename, raytune):
    wb = load_workbook(settingsfilename)
    sheets = wb.sheetnames
    configs_dict = {}
    for sheet in sheets:
        if sheet.lower() == "plans":
            # Read all plans at once and spread them into plan1, plan2, ...
            pass
        else:
            # Old behavior for all other sheets
            configs_dict[sheet] = load_config_from_worksheet(settingsfilename, sheet, raytune)

    return configs_dict
    # sheets.remove("metadata")


def load_metadata(settingsfilename):
    df = pd.read_excel(settingsfilename, sheet_name="metadata", index_col=None)
    return df



def parse_excel_cell(cell_val):
    if isinstance(cell_val, (int, float)):
        return cell_val
    if isinstance(cell_val, str):


        if any(pattern in cell_val for pattern in ("[", "{", "(")):
            try:
                cell_val = ast.literal_eval(cell_val)
                return cell_val
            except:
                return cell_val
            return
        else:
            try:
                return ast.literal_eval(cell_val)
            except:
                return cell_val
    else:
        return cell_val


def parse_excel_cell(cell_val):
    if isinstance(cell_val, (int, float)):
        return _to_py(cell_val)
    if isinstance(cell_val, str):
        if any(p in cell_val for p in ("[", "{", "(")):
            try:
                return _to_py(ast.literal_eval(cell_val))
            except:
                return cell_val
        else:
            try:
                return _to_py(ast.literal_eval(cell_val))
            except:
                return cell_val
    return _to_py(cell_val)
   
def parse_neptune_dict(dic: dict):
    # fixes lists of int appearing in strings
    for kk, vv in dic.items():
        vv = parse_excel_cell(vv)
        dic.update({kk: vv})
    return dic


# %%
if __name__ == "__main__":

# %%
#SECTION:-------------------- setup--------------------------------------------------------------------------------------

    from fran.managers import Project

    P = Project(project_title="litstmp")
    project= P

    C = ConfigMaker(P, raytune=False, configuration_filename=None)
    C.setup(plan_train=4)
    conf= C.configs
    pp( conf['plan_train'])
    
 
# %%
    settingsfilename = "/home/ub/code/fran/configurations/experiment_configs_totalseg.xlsx"
    engine="pd"
    sheet_name="dataset_params"

    if engine == "pd":
        df = pd.read_excel(settingsfilename, sheet_name=sheet_name)
    elif engine == "openpyxl":  # THIS IS NOT PROPERLY IMPLEMENTED AND HAS BUGS
        wb = load_workbook(settingsfilename)
        sheet = wb[sheet_name]
        df = pd.DataFrame(sheet.values)

    else:
        raise NotImplementedError
# %%
    config = {}
    rr = df.iloc[1]
    for row in df.iterrows():
        # rr = row[1]
        rr = check_bool(rr)
        var_type = rr["tune_type"]
        key = rr["var_name"]
        if sheet_name == "transform_factors":
            if raytune == False or rr["tune"] == False:
                val = parse_excel_cell(rr["manual_value"])
                prob = rr["manual_p"]
            else:
                if rr["tune_type"] == "double_range":
                    val_lower, val_upper = parse_excel_cell(rr["tune_value"])
                    val_lower = tune.uniform(lower=val_lower[0], upper=val_lower[1])
                    val_upper = tune.uniform(lower=val_upper[0], upper=val_upper[1])
                    val = [val_lower, val_upper]

                    prob = parse_excel_cell(rr["tune_p"])
                    prob = tune.uniform(lower=prob[0], upper=prob[1])
            config.update({key: [val, prob]})
        else:
            if raytune == False or rr["tune"] == False:
                val = parse_excel_cell(rr["manual_value"])
                config.update({key: parse_excel_cell(val)})
            else:
                if "spec" in var_type:
                    pass
                else:
                    val = parse_excel_cell(rr["tune_value"])
                    if "_" in var_type:
                        tr()
                        vals = parse_excel_cell(rr["tune_value"])
                        var_type.split("_")
                        tune_fnc = resolve_tune_fnc(var_type)
                        val_sample = [tune_fnc(val[0], val[1]) for val in vals]
                    elif (
                        var_type == "randint"
                        or var_type == "loguniform"
                        or var_type == "uniform"
                    ):
                        tune_fnc = resolve_tune_fnc(var_type)
                        val_sample = tune_fnc(val[0], val[1])
                    elif rr["tune_type"] == "double_range":
                        val_lower, val_upper = parse_excel_cell(rr["tune_value"])
                        val_lower = tune.uniform(lower=val_lower[0], upper=val_lower[1])
                        val_upper = tune.uniform(lower=val_upper[0], upper=val_upper[1])
                        val_sample = [val_lower, val_upper]

                    elif rr["tune_type"] == "choice":
                        val_sample = tune.choice(val)
                    else:
                        tune_fnc = resolve_tune_fnc(var_type)
                        if tune_fnc.__name__[0] == "q":
                            val.append(rr["quant"])
                        val_sample = tune_fnc(*val)
                    config.update({key: val_sample})

# %%

    labels_from_remapping = compute_out_labels(C.configs["plan_train"],C.project.global_properties)
# %%
